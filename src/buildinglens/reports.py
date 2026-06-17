"""Client report generation (Excel and PDF) for a single building.

Reuses the populated database and the existing scoring. The executive summary is
written by the LLM and grounded in the building's actual defects; in mock mode it
falls back to a templated summary so reports still generate offline.
"""

from __future__ import annotations

import io
import sqlite3

from .llm import LLMClient, get_llm
from .scoring import building_risk_breakdown

_SEV_ORDER = "CASE severity WHEN 'critical' THEN 0 WHEN 'major' THEN 1 ELSE 2 END"


def _building(conn: sqlite3.Connection, building_id: int) -> dict:
    row = conn.execute("SELECT * FROM buildings WHERE id = ?", (building_id,)).fetchone()
    if row is None:
        raise ValueError(f"No building with id {building_id}")
    return dict(row)


def _defects(conn: sqlite3.Connection, building_id: int) -> list[dict]:
    rows = conn.execute(
        f"SELECT discipline, element, description, location, severity, citation "
        f"FROM defects WHERE building_id = ? ORDER BY {_SEV_ORDER}, discipline",
        (building_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def executive_summary(
    conn: sqlite3.Connection, building_id: int, client: LLMClient | None = None
) -> str:
    """A short executive summary grounded in the building's defects.

    Uses the configured LLM. Falls back to a deterministic templated summary when
    no key is set (mock mode) so reports still generate offline.
    """
    building = _building(conn, building_id)
    defects = _defects(conn, building_id)
    breakdown = building_risk_breakdown(conn, building_id)

    if not defects:
        return (
            "No defects were extracted for this building. Run the extraction step "
            "(make extract) to populate the report."
        )

    client = client or get_llm()
    listing = "\n".join(
        f"- [{d['severity']}] {d['discipline']} / {d['element']}: {d['description']} ({d['location']})"
        for d in defects
    )
    system = (
        "You are a senior building-inspection analyst writing a short executive "
        "summary for an asset manager or insurer. Be factual and concise, and do not "
        "invent anything beyond the provided defects."
    )
    prompt = (
        f"Building: {building['name']}, {building['address']}.\n"
        f"Risk score: {breakdown['risk_score']:.0f}/100. Defects: "
        f"{breakdown['critical']} critical, {breakdown['major']} major, {breakdown['minor']} minor.\n\n"
        f"Defects:\n{listing}\n\n"
        "Write a 4 to 6 sentence executive summary: overall condition, the most "
        "pressing risks, and a recommended next step."
    )
    try:
        text = client.complete(prompt, system=system, max_tokens=600).strip()
    except Exception:
        text = ""

    if not text or text.startswith("[mock]"):
        disciplines = len({d["discipline"] for d in defects})
        text = (
            f"{building['name']} carries a risk score of {breakdown['risk_score']:.0f} out of 100, "
            f"with {breakdown['critical']} critical, {breakdown['major']} major and "
            f"{breakdown['minor']} minor defects across {disciplines} disciplines. "
            "The critical items should be addressed first. A full re-inspection is "
            "recommended within 12 months."
        )
    return text


def build_excel_report(
    conn: sqlite3.Connection, building_id: int, client: LLMClient | None = None
) -> bytes:
    """Return an .xlsx report: a summary sheet and a defects sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    building = _building(conn, building_id)
    defects = _defects(conn, building_id)
    breakdown = building_risk_breakdown(conn, building_id)
    summary = executive_summary(conn, building_id, client=client)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "BuildingLens inspection report"
    ws["A1"].font = Font(bold=True, size=14, color="1A3A5C")

    rows = [
        ("Building", building.get("name")),
        ("Address", building.get("address")),
        ("Source", building.get("source")),
        ("Risk score", f"{breakdown['risk_score']:.0f} / 100"),
        ("Critical defects", breakdown["critical"]),
        ("Major defects", breakdown["major"]),
        ("Minor defects", breakdown["minor"]),
    ]
    r = 3
    for key, value in rows:
        ws.cell(row=r, column=1, value=key).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Executive summary").font = Font(bold=True)
    r += 1
    cell = ws.cell(row=r, column=1, value=summary)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    ws2 = wb.create_sheet("Defects")
    headers = ["Discipline", "Element", "Description", "Location", "Severity", "Citation"]
    for c, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=c, value=h)
    for i, d in enumerate(defects, start=2):
        ws2.cell(row=i, column=1, value=d["discipline"])
        ws2.cell(row=i, column=2, value=d["element"])
        ws2.cell(row=i, column=3, value=d["description"])
        ws2.cell(row=i, column=4, value=d["location"])
        ws2.cell(row=i, column=5, value=d["severity"])
        ws2.cell(row=i, column=6, value=d["citation"])
    for c, w in enumerate([18, 22, 50, 24, 12, 40], start=1):
        ws2.column_dimensions[chr(64 + c)].width = w

    # Real Excel table over the defects, plus conditional formatting on the
    # severity column: minor = green, major = amber, critical = red.
    last_row = len(defects) + 1
    if last_row >= 2:
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.worksheet.table import Table, TableStyleInfo

        table = Table(displayName="Defects", ref=f"A1:F{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        ws2.add_table(table)

        sev_range = f"E2:E{last_row}"
        rules = {
            "minor": PatternFill("solid", fgColor="D1FAE5"),
            "major": PatternFill("solid", fgColor="FEF3C7"),
            "critical": PatternFill("solid", fgColor="FEE2E2"),
        }
        for value, fill in rules.items():
            ws2.conditional_formatting.add(
                sev_range,
                CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill),
            )

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_pdf_report(
    conn: sqlite3.Connection, building_id: int, client: LLMClient | None = None
) -> bytes:
    """Return a .pdf report: header, KPIs, executive summary, and defect table."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    building = _building(conn, building_id)
    defects = _defects(conn, building_id)
    breakdown = building_risk_breakdown(conn, building_id)
    summary = executive_summary(conn, building_id, client=client)

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=A4, topMargin=20 * mm, bottomMargin=18 * mm,
        leftMargin=18 * mm, rightMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=colors.HexColor("#1a3a5c"), fontSize=16)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#2e6da4"), fontSize=11)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9, leading=13)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=7, leading=9)

    story: list = []
    story.append(Paragraph("BuildingLens inspection report", h1))
    story.append(Paragraph(f"{building['name']} - {building['address']}", body))
    story.append(Spacer(1, 4 * mm))

    meta = [
        ["Risk score", f"{breakdown['risk_score']:.0f} / 100"],
        ["Critical", breakdown["critical"]],
        ["Major", breakdown["major"]],
        ["Minor", breakdown["minor"]],
    ]
    mt = Table(meta, colWidths=[40 * mm, 40 * mm])
    mt.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    story.append(mt)
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("Executive summary", h2))
    story.append(Paragraph(summary, body))
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("Defects", h2))
    data = [["Discipline", "Element", "Description", "Location", "Sev."]]
    for d in defects:
        data.append([
            Paragraph(d["discipline"] or "", small),
            Paragraph(d["element"] or "", small),
            Paragraph(d["description"] or "", small),
            Paragraph(d["location"] or "", small),
            Paragraph(d["severity"] or "", small),
        ])
    dt = Table(data, colWidths=[26 * mm, 28 * mm, 70 * mm, 26 * mm, 16 * mm], repeatRows=1)
    style_cmds = [
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dce8f5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    # Colour the severity cell per row: minor = green, major = amber, critical = red.
    sev_bg = {
        "minor": colors.HexColor("#d1fae5"),
        "major": colors.HexColor("#fef3c7"),
        "critical": colors.HexColor("#fee2e2"),
    }
    for ridx, d in enumerate(defects, start=1):
        bg = sev_bg.get(d["severity"])
        if bg is not None:
            style_cmds.append(("BACKGROUND", (4, ridx), (4, ridx), bg))
    dt.setStyle(TableStyle(style_cmds))
    story.append(dt)

    doc.build(story)
    return bio.getvalue()
